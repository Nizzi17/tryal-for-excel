# app.py
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import pandas as pd
import os 
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'data/staff_feedback.xlsx' 
DIVISIONS = {
    'FINANCIAL DERIVATIVES DIVISION': {'header_row': 1, 'data_start_row': 3},
    'COMMODITIES EXCHANGES AND PRODUCTS DIVISION': {'header_row': 101, 'data_start_row': 103},
    'RISK MANAGEMENT DIVISION': {'header_row': 201, 'data_start_row': 203} 
}

headers = [
    'ID', 'Activity', 'Division', 'Start Date', 'Date of Last Update', 'Name', 'Work Done', 
    'Status', 'Recommendation', 'Approval from ECOP (if any)'
]

def get_current_week_sheet_name():
    """Returns sheet name in format 'YYYY-WW' based on current date"""
    today = datetime.now()
    year, week_num, _ = today.isocalendar()
    return f"{year}-W{week_num:02d}"

def init_excel():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    current_sheet_name = get_current_week_sheet_name()
    
    if not os.path.isfile(EXCEL_FILE): 
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            pd.DataFrame().to_excel(writer, sheet_name=current_sheet_name, index=False)
        
        format_sheet(current_sheet_name)
    else:
        wb = load_workbook(EXCEL_FILE)
        if current_sheet_name not in wb.sheetnames:
            wb.create_sheet(current_sheet_name)
            wb.save(EXCEL_FILE)
            format_sheet(current_sheet_name)
        wb.close()

def format_sheet(sheet_name):
    """Format a sheet with the required structure"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['C'].hidden = True
    
    for col in ['B', 'F', 'G', 'I', 'J']:
        ws.column_dimensions[col].width = 48.00

    for col in ['D', 'E', 'H']:
        ws.column_dimensions[col].width = 18.00
    
    header_font = Font(bold=True)
    division_header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    column_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for division, positions in DIVISIONS.items():
        ws.merge_cells(start_row=positions['header_row'], start_column=1, 
                      end_row=positions['header_row'], end_column=len(headers))
        cell = ws.cell(row=positions['header_row'], column=1, value=division)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True, color='FFFFFF', size=22)  
        cell.fill = division_header_fill  
  
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=positions['header_row'] + 1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = column_header_fill 
    
    wb.save(EXCEL_FILE)

def get_next_available_row(sheet_name, division):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    data_start_row = DIVISIONS[division]['data_start_row']
    current_row = data_start_row
    
    while ws.cell(row=current_row, column=1).value is not None:
        current_row += 1
    
    wb.close()
    return current_row

def save_to_excel(entry):
    current_sheet_name = get_current_week_sheet_name()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[current_sheet_name]
    
    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['C'].hidden = True

    division = entry['Division']
    row = get_next_available_row(current_sheet_name, division)
    
    def insert_line_breaks(text, max_length=60):
        """Insert line breaks every max_length characters, trying to preserve words"""
        if not text or not isinstance(text, str):
            return text
            
        words = text.split(' ')
        lines = []
        current_line = ""
        
        for word in words:
            if len(current_line) + len(word) + 1 <= max_length:  
                current_line += (" " + word) if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
            
        return '\n'.join(lines)
  
    activity = insert_line_breaks(entry['Activity'])
    work_done = insert_line_breaks(entry['Work Done'])
    recommendation = insert_line_breaks(entry['Recommendation'])
    approval = insert_line_breaks(entry['Approval from ECOP (if any)'])
    
    ws.cell(row=row, column=1, value=entry['ID'])
    ws.cell(row=row, column=2, value=activity)
    ws.cell(row=row, column=3, value=division)
    ws.cell(row=row, column=4, value=entry['Start Date'])
    ws.cell(row=row, column=5, value=None)
    ws.cell(row=row, column=6, value=entry['Name'])
    ws.cell(row=row, column=7, value=work_done)
    ws.cell(row=row, column=8, value=entry['Status'])
    ws.cell(row=row, column=9, value=recommendation)
    ws.cell(row=row, column=10, value=approval)
    
    for col in [2, 7, 9, 10]: 
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    
    header_font = Font(bold=True)
    column_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for division, positions in DIVISIONS.items():
        cell = ws.cell(row=positions['header_row'], column=1)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=positions['header_row'] + 1, column=col_num)
            cell.font = header_font
            cell.fill = column_header_fill
    
    wb.save(EXCEL_FILE)
    wb.close()

def read_all_entries():
    """Read entries from all sheets"""
    wb = load_workbook(EXCEL_FILE)
    entries = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for division, positions in DIVISIONS.items():
            current_row = positions['data_start_row']
            
            while ws.cell(row=current_row, column=1).value is not None:
                entry = {
                    'ID': ws.cell(row=current_row, column=1).value,
                    'Activity': ws.cell(row=current_row, column=2).value,
                    'Division': ws.cell(row=current_row, column=3).value,
                    'Start Date': ws.cell(row=current_row, column=4).value,
                    'Last Update': ws.cell(row=current_row, column=5).value,
                    'Name': ws.cell(row=current_row, column=6).value,
                    'Work Done': ws.cell(row=current_row, column=7).value,
                    'Status': ws.cell(row=current_row, column=8).value,
                    'Recommendation': ws.cell(row=current_row, column=9).value,
                    'Approval from ECOP (if any)': ws.cell(row=current_row, column=10).value,
                    'Week': sheet_name 
                }
                entries.append(entry)
                current_row += 1
    
    wb.close()
    return entries

def update_entry(entry_id, updated_data):
    """Search through all sheets to find and update the entry"""
    wb = load_workbook(EXCEL_FILE)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for division, positions in DIVISIONS.items():
            current_row = positions['data_start_row']
            
            while ws.cell(row=current_row, column=1).value is not None:
                if ws.cell(row=current_row, column=1).value == entry_id:
                    
                    if 'last_update' in updated_data and updated_data['last_update']:
                        ws.cell(row=current_row, column=5, value=updated_data['last_update'])

                    ws.cell(row=current_row, column=2, value=updated_data['Activity'])
                    ws.cell(row=current_row, column=4, value=updated_data['Start Date'])
                    ws.cell(row=current_row, column=7, value=updated_data['Work Done'])
                    ws.cell(row=current_row, column=8, value=updated_data['Status'])
                    ws.cell(row=current_row, column=9, value=updated_data['Recommendation'])
                    ws.cell(row=current_row, column=10, value=updated_data['Approval from ECOP (if any)'])
                    
                    wb.save(EXCEL_FILE)
                    wb.close()
                    return True
                current_row += 1
    wb.close()
    return False

@app.route('/', methods=['GET', 'POST'])
def index():
    init_excel()
    if request.method == 'POST':
        names = request.form['name']
        division = request.form['division']
        comments = request.form['comment']

        work_done_list = request.form.getlist('work_done[]')
        date_list = request.form.getlist('date[]')
        status_list = request.form.getlist('status[]')
        Activity_list = request.form.getlist('Activity[]')
        recommendation_list = request.form.getlist('recommendation[]')

        all_entries = read_all_entries()
        existing_max_id = max([entry['ID'] for entry in all_entries]) if all_entries else 0

        for i in range(len(work_done_list)):
            entry = {
                'ID': existing_max_id + i + 1,
                'Name': names,
                'Division': division,
                'Work Done': work_done_list[i],
                'Start Date': date_list[i],
                'Status': status_list[i],
                'Activity': Activity_list[i],
                'Recommendation': recommendation_list[i],
                'Approval from ECOP (if any)': comments
            }
            save_to_excel(entry)

        return redirect(url_for('submissions'))

    return render_template('form_multi.html')

@app.route('/submissions')
def submissions():
    init_excel()
    entries = read_all_entries()
    query = request.args.get('q', '').strip().lower()

    if query:
        entries = [entry for entry in entries 
                  if query in str(entry['Name']).lower() or 
                     query in str(entry['Division']).lower() or
                     query in str(entry['Week']).lower()]

    return render_template('submissions.html', data=entries)

@app.route('/download')
def download():
    try:
        if not os.path.exists(EXCEL_FILE):
            return "Report file not found. Please submit some data first.", 404
        
        directory = os.path.abspath(os.path.dirname(EXCEL_FILE))
        filename = os.path.basename(EXCEL_FILE)
        current_sheet = get_current_week_sheet_name()
        
        return send_from_directory(
            directory=directory,
            path=filename,
            as_attachment=True,
            download_name=f"DRMD_Weekly_Report_{current_sheet}.xlsx"
        )
    except Exception as e:
        return f"Download error: {str(e)}", 500

@app.route('/edit/<int:entry_id>', methods=['GET', 'POST'])
def edit(entry_id):
    init_excel()
    entries = read_all_entries()
    entry = next((e for e in entries if e['ID'] == entry_id), None)

    if not entry:
        return "Entry not found.", 404

    if request.method == 'POST':
        updated_data = {
            'Work Done': request.form['work_done'],
            'Start Date': request.form['date'],
            'Status': request.form['status'],
            'Activity': request.form['Activity'],
            'Recommendation': request.form['recommendation'],
            'Approval from ECOP (if any)': request.form['comment'],
            'last_update': request.form.get('last_update')
        }
        update_entry(entry_id, updated_data)
        return redirect(url_for('submissions'))

    return render_template('edit.html', entry=entry)
    
if __name__ == '__main__':
    init_excel()
    app.run(debug=True)